# imports the io library
import io
import re
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


# INPUT: txt file of magic cards
# OUTPUT: a list with each card's name repeated the number of times that card is in the deck
def file_to_list(filename):
    # type: (object) -> object
    f = open(filename, "r")
    # list to hold temporary list of cards
    temp = []
    for line in f:
        temp.append(line[:(len(line) - 1)])
    file.close(f)
    return temp


# INPUT: The list of cards, and the name of the deck
# SIDE EFFECTS: Adds a new sheet for a deck to the Spreadsheet
def add_deck(card_list, deck_name):
    # creates a Card Library excel doc, and if one deoesn't exist. Makes a new one
    try:
        wb = load_workbook("Card Library.xlsx")
    # catches specific openpyxl exception to cannot find excel book
    except openpyxl.shared.exc.InvalidFileException:
        wb = Workbook()
    # checks to make sure deck doesn't already exist
    try:
        ws = wb[deck_name]
    # catches error for that worksheet not existing
    except KeyError:
        ws = wb.create_sheet(0, deck_name)
    # Deck header...
    ws['A1'] = "Qty"
    ws['B1'] = "Name"
    # sets parameters to then iterate through card list to put all in excel doc
    row = 2
    for card in card_list:
        # because list seems to go on a little long...past the end of the cards
        if card:
            # gets the number of copies this card is in the deck
            pos = 'A' + str(row)
            num = re.findall(r'\d', card)
            if num:
                ws[pos] = num[0]
            else:
                ws[pos] = 1
            # attaches the name of the card to the number
            pos = 'B' + str(row)
            ws[pos] = card[2:(len(card))]
        row += 1
    wb.save('Card Library.xlsx')


def update_card_library():
    deck_file_name = raw_input("What's the file name of the deck?: ")
    deck_name = raw_input("And what do you want to call your deck?: ")
    try:
        add_deck(file_to_list(deck_file_name), deck_name)
        print "deck added successfully!"
    except Exception:
        print "Oops! That deck was invalid! Please check the supposed filename of this deck!"


def load_library():
    # creates a Card Library excel doc, and if one deoesn't exist. Makes a new one
    try:
        wb = load_workbook("Card Library.xlsx")
        # for sheet in wb.worksheets:

    # catches specific openpyxl exception to cannot find excel book
    except openpyxl.shared.exc.InvalidFileException:
        print "You don't have a card library yet! Try adding some decks!"


#############################################################################
# script
# loads initial library
try:
    wb = load_workbook("Card Library.xlsx")
    print "Previous library loaded!"
# catches specific openpyxl exception to cannot find excel book
except openpyxl.shared.exc.InvalidFileException:
    wb = Workbook()
    print "New library created!"
    # add the code for "Total" n stuff...
wb.save('Card Library.xlsx')
# UI
while True:
    response = raw_input('What would you like to do?: ').lower()
    if response == "add a deck":
        update_card_library()
    elif response == "modify a deck":
        print "If this deck cannot be found in the Card Library, it will be added automatically!"
        update_card_library()
    elif response == "exit":
        break
    elif response == "help":
        print "Commands available in this program are: add a deck, modify a deck, exit"
    elif response == "load library":
        load_library()
    else:
        print "Could not recognize that command! Type 'help' if stuck!"

# insert other if conditions here...
print "test complete!"
